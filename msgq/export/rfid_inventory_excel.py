"""Exportacion del reporte semanal de instalacion de tags RFID a Excel.

`export_weekly_report(report_df, path)` genera el archivo con el esquema EXACTO
del reporte 'Inventory Tag Installed *.xlsx' que se entrega cada semana:

    TYPE | DATE | ID | Tag | Cost Center | Department | Product

Mismo estilo visual del proyecto hermano (`Inventory_Equipment/export/excel.py`):
encabezado azul corporativo, cuerpo Calibri 10, bordes finos y DATE formateado
como DD/MM/YYYY. La diferencia: DATE es la fecha REAL del cambio (changedAt del
log de auditoria), no la fecha del inventario.

El analisis completo (multi-hoja) reutiliza `export_sheets` (equipment_excel).
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from msgq import config
from msgq.i18n import t, tr_value

_BLUE_FILL = PatternFill("solid", start_color="1F4E78")
_BLUE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=10)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_LEFT = Alignment(horizontal="left", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")


def _cell(v):
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, str):
        return tr_value(v)
    return v


def export_weekly_report(report_df: pd.DataFrame, path: str) -> None:
    """Escribe el reporte semanal (una hoja 'Tag Installed') en `path`."""
    cols = config.WEEKLY_REPORT_COLS
    if report_df is None or report_df.empty:
        df = pd.DataFrame(columns=cols)
    else:
        df = report_df[[c for c in cols if c in report_df.columns]].copy()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tag Installed"

    ws.append([t(str(c)) for c in cols])            # encabezados traducidos
    for _, row in df.iterrows():
        ws.append([_cell(row.get(c)) for c in cols])

    # Encabezado azul.
    for cell in ws[1]:
        cell.fill = _BLUE_FILL
        cell.font = _BLUE_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Cuerpo + formato de fecha en la columna DATE.
    date_idx = cols.index("DATE") + 1 if "DATE" in cols else None
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = _CENTER if cell.column == date_idx else _LEFT
            if cell.column == date_idx and cell.value is not None:
                cell.number_format = "DD/MM/YYYY"

    # Ancho automatico.
    for col in ws.columns:
        best = 10
        for cell in col:
            if cell.value is not None:
                best = max(best, min(len(str(cell.value)) + 2, 50))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

    wb.save(path)
