"""Exportación del análisis de equipos a un workbook de Excel.

`export_sheets(path, sheets)` vuelca un diccionario {nombre_hoja: DataFrame} a un
.xlsx con encabezado azul corporativo, bordes finos y ancho automático — mismo
estilo visual que `Inventory_Equipment/export/excel.py`.
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from msgq.i18n import t, tr_value

_BLUE_FILL = PatternFill("solid", start_color="1F4E78")
_BLUE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=10)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_LEFT = Alignment(horizontal="left", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    # Excel: max 31 chars, sin []:*?/\
    clean = "".join(c for c in str(name) if c not in '[]:*?/\\')[:31] or "Hoja"
    base, i = clean, 1
    while clean in used:
        suffix = f"_{i}"
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(clean)
    return clean


def _cell_value(v):
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if pd.api.types.is_bool(v):   # incluye numpy.bool_ / pandas boolean
        return tr_value("Si" if v else "No")
    if isinstance(v, str):
        return tr_value(v)   # solo tokens conocidos; los datos reales pasan intactos
    return v


def _write_sheet(ws, df: pd.DataFrame) -> None:
    if df is None or df.empty:
        ws.append([t("(sin dato)")])
        return
    ws.append([t(str(c)) for c in df.columns])   # encabezados traducidos
    for _, row in df.iterrows():
        ws.append([_cell_value(v) for v in row])
    for cell in ws[1]:
        cell.fill = _BLUE_FILL
        cell.font = _BLUE_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = _LEFT
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
    # Ancho automatico.
    for col in ws.columns:
        best = 10
        for cell in col:
            if cell.value is not None:
                best = max(best, min(len(str(cell.value)) + 2, 50))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best


def export_sheets(path: str, sheets: dict[str, pd.DataFrame]) -> None:
    """Escribe cada DataFrame en su propia hoja del workbook."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    used: set[str] = set()
    for name, df in sheets.items():
        ws = wb.create_sheet(title=_safe_sheet_name(t(name), used))
        _write_sheet(ws, df)
    if not wb.sheetnames:
        wb.create_sheet("Vacio")
    wb.save(path)
