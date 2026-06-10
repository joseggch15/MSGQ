# -*- coding: utf-8 -*-
"""Pruebas del reporte 'Dispensas por Equipo' (core + export PDF/Excel).

Misma metodologia E2E del proyecto: datos en la forma real del pipeline
(DataFrames con el esquema canonico), SQLite real cuando aplica y archivos de
salida reales en un directorio temporal — sin mocks.

Ejecutar:   pytest tests/test_dispense_report.py -v
o:          python tests/test_dispense_report.py
"""
import os
import shutil
import sys
import tempfile
from datetime import datetime

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from msgq.core import dispense_report as dr


def _movements() -> pd.DataFrame:
    """Despachos sinteticos en el esquema canonico de la replica."""
    rows = [
        # EQ1 (limite real 80 L): 2 normales + 1 over.
        ("M1", "DISPENSE", "2026-01-10T08:00:00", "EQ1", "Toyota Hilux", "Diesel", 50.0),
        ("M2", "DISPENSE", "2026-02-10T08:00:00", "EQ1", "Toyota Hilux", "Diesel", 80.0),
        ("M3", "DISPENSE", "2026-03-10T08:00:00", "EQ1", "Toyota Hilux", "Diesel", 95.0),
        # EQ2 (sin limite, categoria 'Contractor Excavators' -> respaldo 7450):
        # 1 normal + 1 over.
        ("M4", "DISPENSE", "2026-01-15T09:00:00", "EQ2", "Excavadora 374", "Diesel", 7000.0),
        ("M5", "DISPENSE", "2026-02-15T09:00:00", "EQ2", "Excavadora 374", "Diesel", 7500.0),
        # EQ3 (sin limite y sin categoria mapeada): todo Normal por definicion.
        ("M6", "DISPENSE", "2026-01-20T10:00:00", "EQ3", "Bomba de agua", "Diesel", 999.0),
        # Fuera de rango (2025): no debe contar.
        ("M7", "DISPENSE", "2025-06-01T10:00:00", "EQ1", "Toyota Hilux", "Diesel", 999.0),
        # No-DISPENSE: no debe contar.
        ("M8", "DELIVERY", "2026-01-25T10:00:00", "EQ1", "Toyota Hilux", "Diesel", 5000.0),
    ]
    df = pd.DataFrame(rows, columns=["id", "kind", "record_collected_at",
                                     "equipment_id", "equipment_description",
                                     "product", "volume"])
    df["updated_at"] = df["record_collected_at"]
    df["field_user"] = "operador1"
    df["tank"] = "TK-1"
    return df


def _equipment() -> pd.DataFrame:
    return pd.DataFrame({
        "equipment_id": ["EQ1", "EQ2", "EQ3"],
        "description": ["Toyota Hilux LV", "CAT 374 Excavator", "Bomba portatil"],
        "category": ["Light Vehicle", "Contractor Excavators", "Pumps"],
        "group": ["Newmont", "Haukes", "Newmont"],
        "department": ["Mina", "Mina", "Servicios"],
        "make": ["Toyota", "CAT", "ACME"],
        "cost_centre": ["CC1", "CC2", "CC3"],
    })


def _limits() -> pd.DataFrame:
    # Solo EQ1 tiene limite real (80 L para Diesel).
    return pd.DataFrame({
        "id": ["L1"], "equipment_id": ["EQ1"], "internal_id": ["1"],
        "product": ["Diesel"], "product_code": ["D"], "sfl": [80.0],
    })


_FROM = pd.Timestamp("2026-01-01")
_TO = pd.Timestamp("2026-12-31")


def test_dataset_classification_and_sfl_sources():
    ds = dr.build_dataset(_movements(), _equipment(), _limits(), _FROM, _TO)
    assert len(ds) == 6                      # 6 despachos 2026 (sin delivery ni 2025)
    eq1 = ds[ds["equipment_id"] == "EQ1"]
    # volume <= SFL -> Normal (incluye el igual a 80); > SFL -> Over.
    assert list(eq1.sort_values("date")["clase"]) == ["Normal", "Normal", "Over SFL"]
    assert (eq1["sfl"] == 80.0).all()
    assert (eq1["sfl_source"] == dr.SFL_SOURCE_LIMIT).all()
    # La descripcion viene del MAESTRO (merge), no del movimiento.
    assert (eq1["description"] == "Toyota Hilux LV").all()

    eq2 = ds[ds["equipment_id"] == "EQ2"]
    assert (eq2["sfl"] == 7450.0).all()      # respaldo por categoria (config)
    assert (eq2["sfl_source"] == dr.SFL_SOURCE_FALLBACK).all()
    assert list(eq2.sort_values("date")["clase"]) == ["Normal", "Over SFL"]

    eq3 = ds[ds["equipment_id"] == "EQ3"]
    assert eq3["sfl"].isna().all()           # sin limite ni mapeo -> N/D
    assert (eq3["sfl_source"] == dr.SFL_SOURCE_NONE).all()
    assert (eq3["clase"] == "Normal").all()  # no se excede un SFL desconocido
    print("OK  test_dataset_classification_and_sfl_sources")


def test_summaries_and_scope_filters():
    ds = dr.build_dataset(_movements(), _equipment(), _limits(), _FROM, _TO)
    res = dr.equipment_summary(ds)
    r1 = res[res["equipment_id"] == "EQ1"].iloc[0]
    assert r1["Despachos"] == 3 and r1["Normal"] == 2 and r1["Over SFL"] == 1
    assert abs(r1["% Over"] - 33.33) < 0.01

    by_cat = dr.dimension_summary(ds, "category", "Categoría")
    lv = by_cat[by_cat["Categoría"] == "Light Vehicle"].iloc[0]
    assert lv["Equipos"] == 1 and lv["Despachos"] == 3 and lv["Over SFL"] == 1

    # Alcances: ids especificos y dimension/valor.
    only_eq2 = dr.filter_scope(ds, equipment_ids=["EQ2"])
    assert set(only_eq2["equipment_id"]) == {"EQ2"}
    by_group = dr.filter_scope(ds, dimension="group", value="Newmont")
    assert set(by_group["equipment_id"]) == {"EQ1", "EQ3"}

    k = dr.overall_kpis(ds)
    assert k["Equipos"] == 3 and k["Despachos"] == 6
    assert k["Normal"] == 4 and k["Over SFL"] == 2
    print("OK  test_summaries_and_scope_filters")


def test_export_pdf_and_excel_files():
    ds = dr.build_dataset(_movements(), _equipment(), _limits(), _FROM, _TO)
    from msgq.export.dispense_report import export_excel, export_pdf
    work = tempfile.mkdtemp(prefix="msgq_report_")
    try:
        pdf_path = os.path.join(work, "reporte.pdf")
        xlsx_path = os.path.join(work, "reporte.xlsx")
        pages: list = []
        n = export_pdf(pdf_path, ds, scope_label="Prueba",
                       extra_equipment=[("EQ9", "Equipo sin datos")],
                       progress=lambda p, t: pages.append((p, t)))
        assert n == 1 and pages == [(1, 1)]        # 4 graficas caben en 1 pagina
        assert os.path.getsize(pdf_path) > 5000    # PDF real, no vacio
        with open(pdf_path, "rb") as fh:
            assert fh.read(5) == b"%PDF-"

        export_excel(xlsx_path, ds, scope_label="Prueba")
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        for sheet in ("Resumen", "Resumen por equipo", "Despachos Over SFL",
                      "Por categoría", "Despachos clasificados"):
            assert sheet in wb.sheetnames, f"falta hoja {sheet}"
        ws = wb["Resumen por equipo"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 3                      # un equipo por fila
        wb.close()
    finally:
        shutil.rmtree(work, ignore_errors=True)
    print("OK  test_export_pdf_and_excel_files")


def test_pdf_cancellation_stops_early():
    # 19 equipos -> 3 paginas; cancelar tras la primera deja 1 escrita.
    rows = []
    for i in range(19):
        rows.append((f"M{i}", "DISPENSE", "2026-01-10T08:00:00",
                     f"EQ{i:02d}", f"Equipo {i}", "Diesel", 50.0))
    mv = pd.DataFrame(rows, columns=["id", "kind", "record_collected_at",
                                     "equipment_id", "equipment_description",
                                     "product", "volume"])
    mv["updated_at"] = mv["record_collected_at"]
    ds = dr.build_dataset(mv, None, None, _FROM, _TO)
    from msgq.export.dispense_report import export_pdf
    work = tempfile.mkdtemp(prefix="msgq_report_")
    try:
        path = os.path.join(work, "cancelado.pdf")
        state = {"pages": 0}

        def _progress(p, t):
            state["pages"] = p

        n = export_pdf(path, ds, scope_label="Cancel",
                       progress=_progress, cancel=lambda: state["pages"] >= 1)
        assert n == 1                              # se detuvo tras la pagina 1
    finally:
        shutil.rmtree(work, ignore_errors=True)
    print("OK  test_pdf_cancellation_stops_early")


if __name__ == "__main__":
    tests = [
        test_dataset_classification_and_sfl_sources,
        test_summaries_and_scope_filters,
        test_export_pdf_and_excel_files,
        test_pdf_cancellation_stops_early,
    ]
    failed = 0
    for fn in tests:
        try:
            fn()
        except Exception as exc:  # noqa: BLE001
            failed += 1
            import traceback
            print(f"FALLO  {fn.__name__}: {exc}")
            traceback.print_exc()
    print(f"\n{len(tests) - failed}/{len(tests)} pruebas del reporte superadas.")
    raise SystemExit(1 if failed else 0)
