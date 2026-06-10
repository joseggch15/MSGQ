# -*- coding: utf-8 -*-
"""Pruebas del modulo 'Inventario de Tags RFID' (reporte Inventory Tag Installed).

METODOLOGIA (igual que el resto del ecosistema): sin mocks. Las pruebas
deterministas construyen DataFrames con el esquema real (`config.*_COLS`) y la
prueba smoke ejercita el pipeline completo simulador -> transform -> SQLite real
-> reporte.

Verifica el aporte central del modulo frente a `Inventory_Equipment`: la columna
DATE es la fecha REAL del cambio (changed_at del log), y el enlace tag->equipo es
por VALOR contra el maestro (vacio en remociones).

Ejecutar:   pytest tests/test_rfid_inventory.py -v
o:          python tests/test_rfid_inventory.py
"""
import os
import sys
from datetime import datetime, timedelta

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from msgq import config
from msgq.api import make_source
from msgq.core import rfid_inventory as ri
from msgq.core import transform
from msgq.export import export_sheets, export_weekly_report
from msgq.storage import Database


# ---------------------------------------------------------------------------
# Constructores de datos (esquema real de las tablas)
# ---------------------------------------------------------------------------

def _changes_df(rows: list[dict]) -> pd.DataFrame:
    full = []
    for r in rows:
        d = {c: r.get(c) for c in config.CHANGE_EVENT_COLS}
        d["event_key"] = (f"{r.get('record_type')}:{r.get('record_id')}:"
                          f"{r.get('changed_at')}:{r.get('attribute')}")
        full.append(d)
    df = pd.DataFrame(full, columns=config.CHANGE_EVENT_COLS)
    df["changed_at"] = pd.to_datetime(df["changed_at"])
    return df


def _rfid(changed_at, rid, before, after, who="m.venegas@plgims.com") -> dict:
    event = "create" if before is None else ("destroy" if after is None else "update")
    return {"changed_at": changed_at, "record_type": config.CHANGE_RECORD_RFID,
            "record_id": rid, "event": event, "whodunnit": who,
            "attribute": config.ATTR_RFID, "before": before, "after": after}


def _equipment_df(rows: list[dict]) -> pd.DataFrame:
    base = [{**{c: pd.NA for c in config.EQUIPMENT_COLS}, **r} for r in rows]
    return pd.DataFrame(base, columns=config.EQUIPMENT_COLS)


def _movements_df(rows: list[dict]) -> pd.DataFrame:
    base = [{**{c: pd.NA for c in config.MOVEMENT_COLS}, **r} for r in rows]
    return pd.DataFrame(base, columns=config.MOVEMENT_COLS)


# Maestro y movimientos compartidos por varias pruebas.
def _sample_equipment() -> pd.DataFrame:
    return _equipment_df([
        {"equipment_id": "EQ1", "internal_id": "1", "description": "CAT 785",
         "status": config.STATUS_IN, "group": "Haul Trucks", "category": "Truck",
         "department": "Mining", "cost_centre": "CC100", "rfid": "TAGA"},
        {"equipment_id": "EQ2", "internal_id": "2", "description": "Toyota LC",
         "status": config.STATUS_IN, "group": "Light Vehicles", "category": "LV",
         "department": "Geology", "cost_centre": "CC200", "rfid": "TAGB"},
        {"equipment_id": "EQ3", "internal_id": "3", "description": "Komatsu",
         "status": config.STATUS_IN, "group": "Dozers", "category": "Dozer",
         "department": "Mining", "cost_centre": "CC100", "rfid": pd.NA},
    ])


def _sample_movements() -> pd.DataFrame:
    return _movements_df([
        {"id": "m1", "kind": config.KIND_DISPENSE, "equipment_id": "EQ1", "product": "Diesel"},
        {"id": "m2", "kind": config.KIND_DISPENSE, "equipment_id": "EQ1", "product": "Diesel"},
        {"id": "m3", "kind": config.KIND_DISPENSE, "equipment_id": "EQ2", "product": "Unleaded Gasoline"},
    ])


# ===========================================================================
# 1. Clasificacion + FECHA REAL + enlace por valor
# ===========================================================================

def test_report_classifies_types_and_uses_real_date():
    eq = _sample_equipment()
    mv = _sample_movements()
    d_new = pd.Timestamp("2026-05-26 14:17:33")
    d_rep = pd.Timestamp("2026-05-28 09:42:41")
    d_rem = pd.Timestamp("2026-05-29 08:12:30")
    ch = _changes_df([
        _rfid(d_new, "101", None, "TAGA"),       # alta -> EQ1
        _rfid(d_rep, "102", "OLDTAG", "TAGB"),    # reemplazo -> EQ2
        _rfid(d_rem, "103", "TAGC", None),        # remocion (tag ya no existe)
    ])
    rep = ri.installation_report(ch, eq, mv)

    assert list(rep.columns) == ri._REPORT_COLS
    assert len(rep) == 3
    by_tag = {r["Tag"]: r for _, r in rep.iterrows()}

    new = by_tag["TAGA"]
    assert new["TYPE"] == config.TYPE_NEW
    assert new["ID"] == "EQ1"
    assert new["DATE"] == d_new                      # << FECHA REAL del log
    assert new["Cost Center"] == "CC100"
    assert new["Department"] == "Mining"
    assert new["Product"] == "Diesel"                # inferido de despachos

    rep_row = by_tag["TAGB"]
    assert rep_row["TYPE"] == config.TYPE_REPLACEMENT
    assert rep_row["ID"] == "EQ2"
    assert rep_row["DATE"] == d_rep
    assert rep_row["Product"] == "Unleaded Gasoline"

    rem = by_tag["TAGC"]
    assert rem["TYPE"] == config.TYPE_REMOVAL
    assert rem["DATE"] == d_rem
    # sin historial, una remocion no se puede atribuir -> ID marcado, no en blanco
    assert rem["ID"] == config.UNIDENTIFIED
    assert not ri._identified(rem["ID"])
    print("OK  test_report_classifies_types_and_uses_real_date")


def test_value_join_is_case_insensitive():
    eq = _equipment_df([{"equipment_id": "EQX", "internal_id": "9",
                         "status": config.STATUS_IN, "rfid": "abc123"}])
    ch = _changes_df([_rfid(pd.Timestamp("2026-05-20"), "1", None, "ABC123")])
    rep = ri.installation_report(ch, eq, _movements_df([]))
    assert len(rep) == 1
    assert rep.iloc[0]["ID"] == "EQX"
    print("OK  test_value_join_is_case_insensitive")


def test_date_range_filters_events():
    eq = _sample_equipment()
    old = pd.Timestamp.now() - pd.Timedelta(days=400)
    recent = pd.Timestamp.now() - pd.Timedelta(days=3)
    ch = _changes_df([
        _rfid(old, "1", None, "TAGA"),
        _rfid(recent, "2", None, "TAGB"),
    ])
    full = ri.installation_report(ch, eq, _movements_df([]))
    assert len(full) == 2
    last30 = ri.installation_report(
        ch, eq, _movements_df([]),
        date_from=pd.Timestamp.now() - pd.Timedelta(days=30),
        date_to=pd.Timestamp.now())
    assert len(last30) == 1
    assert last30.iloc[0]["Tag"] == "TAGB"
    print("OK  test_date_range_filters_events")


# ===========================================================================
# 2. KPIs y producto
# ===========================================================================

def test_summary_kpis():
    eq = _sample_equipment()
    ch = _changes_df([
        _rfid(pd.Timestamp("2026-05-01"), "1", None, "TAGA"),
        _rfid(pd.Timestamp("2026-05-02"), "2", "OLD", "TAGB"),
        _rfid(pd.Timestamp("2026-05-03"), "3", "TAGC", None),
    ])
    rep = ri.installation_report(ch, eq, _movements_df([]))
    k = ri.summary_kpis(rep, eq)
    assert k["Nuevas instalaciones"] == 1
    assert k["Reemplazos"] == 1
    assert k["Remociones"] == 1
    assert k["Total con RFID"] == 2          # EQ1, EQ2 (EQ3 sin tag)
    assert k["Total equipos"] == 3
    print("OK  test_summary_kpis")


def test_product_map_uses_mode():
    mv = _movements_df([
        {"id": "a", "kind": config.KIND_DISPENSE, "equipment_id": "EQ1", "product": "Diesel"},
        {"id": "b", "kind": config.KIND_DISPENSE, "equipment_id": "EQ1", "product": "Diesel"},
        {"id": "c", "kind": config.KIND_DISPENSE, "equipment_id": "EQ1", "product": "Unleaded Gasoline"},
    ])
    pm = ri.equipment_product_map(mv)
    assert pm["EQ1"] == "Diesel"
    print("OK  test_product_map_uses_mode")


def _limits_df(rows: list[dict]) -> pd.DataFrame:
    base = [{**{c: pd.NA for c in config.CONSUMPTION_LIMIT_COLS}, **r} for r in rows]
    return pd.DataFrame(base, columns=config.CONSUMPTION_LIMIT_COLS)


def test_product_map_prefers_enabled_products():
    """El producto HABILITADO (consumption_limits, 'Products consumed' de
    AdaptIQ) manda sobre el inferido por despachos, y cubre a los equipos
    recien tagueados SIN despachos (antes la columna Product salia vacia)."""
    mv = _movements_df([
        {"id": "a", "kind": config.KIND_DISPENSE, "equipment_id": "EQ1", "product": "Diesel"},
    ])
    lim = _limits_df([
        # EQ1 con dos productos habilitados -> se unen ordenados.
        {"id": "l1", "equipment_id": "EQ1", "product": "Diesel", "sfl": 400.0},
        {"id": "l2", "equipment_id": "EQ1", "product": "15W40", "sfl": 20.0},
        # EQ9: SIN despachos (equipo nuevo) -> el producto sale del limite.
        {"id": "l3", "equipment_id": "EQ9", "product": "Diesel", "sfl": 400.0},
    ])
    pm = ri.equipment_product_map(mv, lim)
    assert pm["EQ1"] == "15W40, Diesel"
    assert pm["EQ9"] == "Diesel"
    # Sin limites se conserva el comportamiento previo (mas despachado).
    assert ri.equipment_product_map(mv)["EQ1"] == "Diesel"
    print("OK  test_product_map_prefers_enabled_products")


def test_report_product_from_limits_and_duplicate_master_bridge():
    """Caso real C-SE-12: el maestro tiene el equipo DUPLICADO ('C- SE-12' con
    espacio y 'C-SE-12' sin el, mismo tag), el limite de producto cuelga solo
    de la variante sin espacio y el equipo no ha despachado nunca. El reporte
    debe mostrar igualmente el producto asignado (puente por id compacto)."""
    eq = _equipment_df([
        {"equipment_id": "C- SE-12", "internal_id": "10",
         "description": "General Contractor - SEMC", "status": config.STATUS_OUT,
         "cost_centre": "10001519", "department": "SUSTAINING CAPEX",
         "rfid": "56B59209"},
        {"equipment_id": "C-SE-12", "internal_id": "11",
         "description": "General Contractor - SEMC", "status": config.STATUS_OUT,
         "cost_centre": "10001519", "department": "SUSTAINING CAPEX",
         "rfid": "56B59209"},
    ])
    lim = _limits_df([
        {"id": "l1", "equipment_id": "C-SE-12", "product": "Diesel", "sfl": 400.0},
    ])
    ch = _changes_df([
        _rfid(pd.Timestamp("2026-06-09 10:00:00"), "201", None, "56B59209"),
    ])
    rep = ri.installation_report(ch, eq, _movements_df([]), limits=lim)
    assert len(rep) == 1
    row = rep.iloc[0]
    assert row["TYPE"] == config.TYPE_NEW
    assert row["Product"] == "Diesel"        # via limite + puente de duplicado
    print("OK  test_report_product_from_limits_and_duplicate_master_bridge")


# ===========================================================================
# 3. Validaciones
# ===========================================================================

def test_validation_duplicate_tags_in_master():
    eq = _equipment_df([
        {"equipment_id": "EQ1", "internal_id": "1", "status": config.STATUS_IN, "rfid": "DUP"},
        {"equipment_id": "EQ2", "internal_id": "2", "status": config.STATUS_IN, "rfid": "DUP"},
        {"equipment_id": "EQ3", "internal_id": "3", "status": config.STATUS_IN, "rfid": "UNIQ"},
    ])
    dup = ri.find_duplicate_tags(eq)
    assert len(dup) == 2
    assert set(dup["equipment_id"]) == {"EQ1", "EQ2"}
    assert (dup["Equipos con este tag"] == 2).all()
    print("OK  test_validation_duplicate_tags_in_master")


def test_validation_out_of_service_and_incomplete():
    eq = _equipment_df([
        {"equipment_id": "EQOOS", "internal_id": "1", "status": config.STATUS_OUT,
         "department": "Mining", "cost_centre": "CC1", "rfid": "TOOS"},
    ])
    ch = _changes_df([
        _rfid(pd.Timestamp("2026-05-01"), "1", None, "TOOS"),     # OOS con tag
        _rfid(pd.Timestamp("2026-05-02"), "2", None, "GHOST"),    # tag no esta en maestro
    ])
    rep = ri.installation_report(ch, eq, _movements_df([]))
    oos = ri.find_out_of_service(rep)
    assert len(oos) == 1 and oos.iloc[0]["Tag"] == "TOOS"
    inc = ri.find_incomplete_records(rep)          # alta sin equipo (GHOST)
    assert len(inc) == 1 and inc.iloc[0]["Tag"] == "GHOST"
    print("OK  test_validation_out_of_service_and_incomplete")


def test_validation_duplicate_ids_retag():
    # Un equipo con dos tags y dos altas en el periodo -> aparece 2 veces.
    eq = _equipment_df([{"equipment_id": "EQ1", "internal_id": "1",
                         "status": config.STATUS_IN, "rfid": "M1, M2"}])
    ch = _changes_df([
        _rfid(pd.Timestamp("2026-05-01"), "1", None, "M1"),
        _rfid(pd.Timestamp("2026-05-10"), "2", None, "M2"),
    ])
    rep = ri.installation_report(ch, eq, _movements_df([]))
    did = ri.find_duplicate_ids(rep)
    assert len(did) == 2
    assert (did["ID"] == "EQ1").all()
    print("OK  test_validation_duplicate_ids_retag")


def test_groupings():
    eq = _sample_equipment()
    ch = _changes_df([
        _rfid(pd.Timestamp("2026-05-01"), "1", None, "TAGA"),   # Mining / CC100
        _rfid(pd.Timestamp("2026-05-02"), "2", None, "TAGB"),   # Geology / CC200
    ])
    rep = ri.installation_report(ch, eq, _movements_df([]))
    dep = ri.by_department_summary(rep)
    assert set(dep["Departamento"]) == {"Mining", "Geology"}
    assert int(dep["Instalaciones"].sum()) == 2
    bt = ri.by_type_summary(rep)
    assert int(bt["Cantidad"].sum()) == 2
    print("OK  test_groupings")


# ===========================================================================
# 3b. Historial tag->equipo (resuelve tags removidos / reemplazados)
# ===========================================================================

def test_history_fallback_fills_removed_tag():
    # El tag ya NO esta en rfidTags del equipo (se removio), pero el historial
    # recuerda que pertenecio a EQ1 -> el reporte lo resuelve (alta Y baja).
    eq = _equipment_df([
        {"equipment_id": "EQ1", "internal_id": "1", "description": "CAT 785",
         "status": config.STATUS_IN, "department": "Mining", "cost_centre": "CC100",
         "rfid": pd.NA}])
    hist = pd.DataFrame([{"tag": "TGONE", "equipment_id": "EQ1", "internal_id": "1",
                          "last_seen": pd.Timestamp("2026-05-01")}],
                        columns=config.RFID_HISTORY_COLS)
    ch = _changes_df([
        _rfid(pd.Timestamp("2026-05-02"), "1", None, "TGONE"),    # alta
        _rfid(pd.Timestamp("2026-05-09"), "1", "TGONE", None),    # baja
    ])
    rep = ri.installation_report(ch, eq, _movements_df([]), history=hist)
    assert len(rep) == 2
    for _, r in rep.iterrows():
        assert r["ID"] == "EQ1"                  # resuelto por historial
        assert r["Cost Center"] == "CC100"       # atributos desde el maestro actual
    print("OK  test_history_fallback_fills_removed_tag")


def test_unidentified_marker_when_no_history():
    eq = _equipment_df([{"equipment_id": "EQ1", "internal_id": "1",
                         "status": config.STATUS_IN, "rfid": "OTHER"}])
    ch = _changes_df([_rfid(pd.Timestamp("2026-05-02"), "9", None, "MYSTERY")])
    rep = ri.installation_report(ch, eq, _movements_df([]))   # sin history
    assert rep.iloc[0]["ID"] == config.UNIDENTIFIED
    assert len(ri.find_incomplete_records(rep)) == 1          # alta sin equipo
    print("OK  test_unidentified_marker_when_no_history")


def test_rfid_history_assignments_df_and_roundtrip():
    eq = _equipment_df([
        {"equipment_id": "EQ1", "internal_id": "1", "status": config.STATUS_IN,
         "rfid": "tag-a, tag-b"},
        {"equipment_id": "EQ2", "internal_id": "2", "status": config.STATUS_IN,
         "rfid": pd.NA},
    ])
    df = transform.rfid_assignments_df(eq, pd.Timestamp("2026-05-01"))
    assert set(df["tag"]) == {"TAG-A", "TAG-B"}      # mayusculas, separado por coma
    assert (df["equipment_id"] == "EQ1").all()
    db = Database(":memory:")
    db.upsert("rfid_history", df)
    back = db.get_rfid_history()
    db.close()
    assert set(back["tag"]) == {"TAG-A", "TAG-B"}
    print("OK  test_rfid_history_assignments_df_and_roundtrip")


# ===========================================================================
# 4. Smoke E2E con el simulador (pipeline real + SQLite)
# ===========================================================================

def test_e2e_report_via_simulator():
    import asyncio

    src = make_source(config.Settings(demo_mode=True, token="", db_path=":memory:"))
    ch_nodes = asyncio.run(src.fetch_changes(config.CHANGE_RECORD_RFID, None))
    eq_nodes = asyncio.run(src.fetch_equipment(None))
    mv_nodes = asyncio.run(src.fetch_movements(None))
    asyncio.run(src.aclose())

    db = Database(":memory:")
    db.upsert("change_events", transform.change_events_to_df(ch_nodes))
    db.upsert("equipment", transform.equipment_to_df(eq_nodes))
    db.upsert("movements", transform.movements_to_df(mv_nodes))

    changes = db.get_change_events(config.CHANGE_RECORD_RFID)
    eq = db.get_equipment()
    mv = db.read("movements")
    db.close()

    rep = ri.installation_report(changes, eq, mv)
    assert list(rep.columns) == ri._REPORT_COLS
    assert not rep.empty, "el simulador emite altas de RFID -> el reporte no debe ir vacio"
    assert set(rep["TYPE"]).issubset(
        {config.TYPE_NEW, config.TYPE_REPLACEMENT, config.TYPE_REMOVAL})
    assert pd.api.types.is_datetime64_any_dtype(rep["DATE"])
    # report_display y validaciones no deben fallar
    assert list(ri.report_display(rep).columns) == ri.DISPLAY_COLS
    _ = ri.validation_summary(rep, eq)
    print("OK  test_e2e_report_via_simulator")


def test_exports_roundtrip(tmp_path=None):
    import tempfile
    from openpyxl import load_workbook

    eq = _sample_equipment()
    ch = _changes_df([
        _rfid(pd.Timestamp("2026-05-01"), "1", None, "TAGA"),
        _rfid(pd.Timestamp("2026-05-02"), "2", "OLD", "TAGB"),
    ])
    rep = ri.installation_report(ch, eq, _sample_movements())

    d = tempfile.mkdtemp()
    weekly = os.path.join(d, "weekly.xlsx")
    export_weekly_report(rep, weekly)
    wb = load_workbook(weekly)
    ws = wb.active
    assert ws.max_row == 1 + len(rep)                 # encabezado + filas
    assert [c.value for c in ws[1]][:2] == ["TYPE", "DATE"]  # esquema semanal

    full = os.path.join(d, "full.xlsx")
    export_sheets(full, {
        "Reporte semanal": ri.report_display(rep),
        "Resumen validacion": ri.validation_summary(rep, eq),
    })
    assert os.path.isfile(full)
    print("OK  test_exports_roundtrip")


if __name__ == "__main__":
    tests = [
        test_report_classifies_types_and_uses_real_date,
        test_value_join_is_case_insensitive,
        test_date_range_filters_events,
        test_summary_kpis,
        test_product_map_uses_mode,
        test_product_map_prefers_enabled_products,
        test_report_product_from_limits_and_duplicate_master_bridge,
        test_validation_duplicate_tags_in_master,
        test_validation_out_of_service_and_incomplete,
        test_validation_duplicate_ids_retag,
        test_groupings,
        test_history_fallback_fills_removed_tag,
        test_unidentified_marker_when_no_history,
        test_rfid_history_assignments_df_and_roundtrip,
        test_e2e_report_via_simulator,
        test_exports_roundtrip,
    ]
    failed = 0
    for tc in tests:
        try:
            tc()
        except Exception as exc:  # noqa: BLE001
            failed += 1
            import traceback
            print(f"FALLO  {tc.__name__}: {exc}")
            traceback.print_exc()
    print(f"\n{len(tests) - failed}/{len(tests)} pruebas superadas.")
    raise SystemExit(1 if failed else 0)
